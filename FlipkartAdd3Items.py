import time
import openpyxl

from selenium.webdriver import Keys, ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


class FlipKart:
    def __init__(self, driver):
        self.book = openpyxl.load_workbook("/Users/preepawar/Documents/excelData.xlsx")
        self.sheet = self.book.active
        self.cell1 = self.sheet.cell(row=2, column=2)
        self.cell2 = self.sheet.cell(row=3, column=2)
        self.cell3 = self.sheet.cell(row=4, column=2)

        # first item details
        self.driver = driver
        self.searchGrocery = (By.XPATH, "//input[@class='_3704LK']")
        self.groceryItem = (By.XPATH, "/html[1]/body[1]/div[1]/div[1]/div[3]/div[1]/div[2]/div[2]/div[1]/div[1]/div["
                                      "1]/a[1]/div[1]/div[1]/div[1]/img[1]")
        self.addItem = (By.XPATH, "//button[@class='_2KpZ6l _2U9uOA _3v1-ww']")

        # second item details
        self.searchMobile = (By.XPATH, "//input[@class='_3704LK']")
        self.mobileItem = (By.XPATH, "(//img[@class='_396cs4'])[1]")

        # Third item details
        self.searchWallets = (By.XPATH, "//input[@class='_3704LK']")
        self.wallet = (By.XPATH, "(//img[@class='_2r_T1I'])[1]")

        # validate Cart items
        self.cart = (By.XPATH, "//div[@class='KK-o3G']")

    def select_firstItem(self):
        grocery = self.driver.find_element(*self.searchGrocery)
        grocery.send_keys(self.cell1.value)
        grocery.send_keys(Keys.ENTER)
        time.sleep(5)
        itemValue = grocery.text
        print(itemValue)
        time.sleep(5)

        current_window = self.driver.current_window_handle
        self.driver.find_element(*self.groceryItem).click()
        time.sleep(2)

        for window_handle in self.driver.window_handles:
            if window_handle != current_window:
                self.driver.switch_to.window(window_handle)
                break

        self.driver.find_element(*self.addItem).click()
        time.sleep(2)

        self.driver.close()

        # switch to previous window
        self.driver.switch_to.window(current_window)
        time.sleep(2)

    def select_secondItem(self):
        mobile = self.driver.find_element(*self.searchMobile)
        mobile.clear()
        mobile.send_keys(self.cell2.value)
        mobile.send_keys(Keys.ENTER)
        time.sleep(5)
        itemValue = mobile.text
        print(itemValue)
        time.sleep(5)
        current_window = self.driver.current_window_handle
        self.driver.find_element(*self.mobileItem).click()
        time.sleep(2)

        for window_handle in self.driver.window_handles:
            if window_handle != current_window:
                self.driver.switch_to.window(window_handle)
                break

        self.driver.find_element(*self.addItem).click()
        time.sleep(2)

        self.driver.close()

        # switch to previous window
        self.driver.switch_to.window(current_window)
        time.sleep(2)

    def select_thirdItem(self):
        wallets = self.driver.find_element(*self.searchWallets)
        wallets.clear()
        wallets.send_keys(self.cell3.value)
        wallets.send_keys(Keys.ENTER)
        time.sleep(5)
        itemValue = wallets.text
        print(itemValue)
        time.sleep(5)
        current_window = self.driver.current_window_handle
        self.driver.find_element(*self.wallet).click()
        time.sleep(3)

        for window_handle in self.driver.window_handles:
            if window_handle != current_window:
                self.driver.switch_to.window(window_handle)
                break

        self.driver.execute_script("window.scrollBy(0, 500);")
        time.sleep(3)

        addItem = self.driver.find_element(*self.addItem)
        # self.driver.execute_script("arguments[0].scrollIntoView();", addItem)
        actions = ActionChains(self.driver)

        # Move to the element and click using ActionChains
        actions.move_to_element(addItem).click().perform()

        time.sleep(2)

        self.driver.close()

        # switch to previous window
        self.driver.switch_to.window(current_window)
        time.sleep(5)

    def validate_CartItems(self):
        self.driver.refresh()
        cartCount = self.driver.find_element(*self.cart).text
        expected = '3'
        assert cartCount in expected
        time.sleep(2)
        self.driver.quit()
